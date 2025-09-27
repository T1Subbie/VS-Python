import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import pandas as pd
from datetime import datetime
import os
import sys
import subprocess
from tkinter import messagebox
from fpdf import FPDF
import glob
import re
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
import traceback

# --- CONFIGURAÇÕES GLOBAIS ---
PASTA_LOGS_EXCEL = "Logs_Excel"
PASTA_ROMANEIOS_PDF = "Romaneios_PDF"
TIPOS_CONTAINER = ["20' DC", "40' DC", "40' HC", "20' REEFER", "40' REEFER", "Outro"]
CONDICOES = ["Cheio", "Vazio"]
NOME_ABA_EXCEL = "Movimentações"
COLUNAS_ORDENADAS = [
    'Data e Hora', 'Status', 'Nº do Contêiner', 'Tipo de Contêiner', 'Condição',
    'Cliente', 'Nº do Lacre', 'Nota Fiscal (NF)', 'Destino', 'Observações',
    'Placa do Veículo', 'Placa Carreta', 'Motorista', 'CPF Motorista', 'Transportadora',
    'Tara', 'Peso Bruto Carga', 'Booking', 'Armador', 'Navio', 'Deadline'
]

# --- UTIL ---
def _base_dir():
    # garante caminho absoluto (evita confusão se rodar de outra pasta)
    return os.path.dirname(os.path.abspath(__file__))

def _abrir_no_sistema(caminho: str):
    try:
        if sys.platform.startswith("win"):
            os.startfile(caminho)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", caminho], check=False)
        else:
            subprocess.run(["xdg-open", caminho], check=False)
    except Exception:
        pass

# --- FUNÇÕES DE VALIDAÇÃO ---
def validar_numero_container(numero_container):
    num_limpo = re.sub(r'[^A-Z0-9]', '', str(numero_container).upper())
    if not re.match(r'^[A-Z]{4}\d{7}$', num_limpo):
        return False
    mapa_valores = {
        '0': 0, '1': 1, '2': 2, '3': 3, '4': 4, '5': 5, '6': 6, '7': 7, '8': 8, '9': 9,
        'A': 10, 'B': 12, 'C': 13, 'D': 14, 'E': 15, 'F': 16, 'G': 17, 'H': 18, 'I': 19,
        'J': 20, 'K': 21, 'L': 23, 'M': 24, 'N': 25, 'O': 26, 'P': 27, 'Q': 28, 'R': 29,
        'S': 30, 'T': 31, 'U': 32, 'V': 34, 'W': 35, 'X': 36, 'Y': 37, 'Z': 38
    }
    soma_ponderada = sum(mapa_valores[num_limpo[i]] * (2**i) for i in range(10))
    digito_calculado = soma_ponderada % 11
    if digito_calculado == 10:
        digito_calculado = 0
    return digito_calculado == int(num_limpo[10])

def validar_cpf(cpf):
    cpf_numeros = ''.join(filter(str.isdigit, cpf))
    if len(cpf_numeros) != 11 or cpf_numeros == cpf_numeros[0] * 11:
        return False
    soma = sum(int(cpf_numeros[i]) * (10 - i) for i in range(9))
    resto = (soma * 10) % 11
    if resto == 10:
        resto = 0
    if resto != int(cpf_numeros[9]):
        return False
    soma = sum(int(cpf_numeros[i]) * (11 - i) for i in range(10))
    resto = (soma * 10) % 11
    if resto == 10:
        resto = 0
    if resto != int(cpf_numeros[10]):
        return False
    return True

# --- ARQUIVOS / PLANILHA ---
def get_caminho_log_diario():
    hoje = datetime.now().strftime('%Y-%m-%d')
    pasta_logs_dia = os.path.join(_base_dir(), PASTA_LOGS_EXCEL, hoje)
    os.makedirs(pasta_logs_dia, exist_ok=True)
    nome_arquivo = f"Log_Diario_{hoje}.xlsx"
    return os.path.join(pasta_logs_dia, nome_arquivo)

def get_pasta_logs_do_dia():
    hoje = datetime.now().strftime('%Y-%m-%d')
    return os.path.join(_base_dir(), PASTA_LOGS_EXCEL, hoje)

def garantir_arquivo_do_dia():
    caminho = get_caminho_log_diario()
    if not os.path.exists(caminho):
        df_vazio = pd.DataFrame(columns=COLUNAS_ORDENADAS)
        salvar_planilha(df_vazio, caminho)
    return caminho

def formatar_planilha(writer, df):
    ws = writer.sheets[NOME_ABA_EXCEL]
    header_font = Font(bold=True, color="FFFFFF", name='Calibri')
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    fill_entrada = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_saida = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    formato_data = "DD/MM/YYYY HH:MM"

    max_row = ws.max_row
    max_col = ws.max_column

    for col_num, column_title in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_num)
        ws.cell(row=1, column=col_num).font = header_font
        ws.cell(row=1, column=col_num).fill = header_fill
        max_length = max(df[column_title].astype(str).map(len).max(), len(column_title)) if not df.empty else len(column_title)
        ws.column_dimensions[col_letter].width = (max_length + 2)
        if column_title == 'Data e Hora':
            for cell in ws[col_letter]:
                if cell.row > 1:
                    cell.number_format = formato_data

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    try:
        col_status_idx = df.columns.get_loc('Status') + 1
        col_status_letter = get_column_letter(col_status_idx)
        data_range_full = f"A2:{get_column_letter(max_col)}{max_row}"
        rule_entrada = FormulaRule(formula=[f'${col_status_letter}2="Entrada"'], fill=fill_entrada)
        rule_saida = FormulaRule(formula=[f'${col_status_letter}2="Saída"'], fill=fill_saida)
        ws.conditional_formatting.add(data_range_full, rule_entrada)
        ws.conditional_formatting.add(data_range_full, rule_saida)
    except KeyError:
        pass

def salvar_planilha(df_para_salvar, caminho_arquivo):
    if 'Data e Hora' in df_para_salvar.columns:
        df_para_salvar['Data e Hora'] = pd.to_datetime(df_para_salvar['Data e Hora'])
    df_para_salvar = df_para_salvar.reindex(columns=COLUNAS_ORDENADAS)
    df_para_salvar = df_para_salvar.sort_values(by='Data e Hora', ascending=False, na_position='last')

    df_para_escrever = df_para_salvar.copy()
    for col in df_para_escrever.columns:
        if col != 'Data e Hora':
            df_para_escrever[col] = df_para_escrever[col].fillna('')

    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl', datetime_format=None) as writer:
        df_para_escrever.to_excel(writer, sheet_name=NOME_ABA_EXCEL, index=False)
        formatar_planilha(writer, df_para_escrever)

def registrar_movimento(status, dados_base=None):
    caminho_arquivo = get_caminho_log_diario()
    novo_df = pd.DataFrame([dados_base])
    try:
        try:
            df_existente = pd.read_excel(caminho_arquivo, sheet_name=NOME_ABA_EXCEL)
            for col in df_existente.columns:
                if col != 'Data e Hora':
                    df_existente[col] = df_existente[col].astype(str)
        except (FileNotFoundError, ValueError):
            df_existente = pd.DataFrame(columns=COLUNAS_ORDENADAS).astype('object')

        df_atualizado = pd.concat([df_existente, novo_df], ignore_index=True)
        salvar_planilha(df_atualizado, caminho_arquivo)
        return True
    except Exception as e:
        traceback.print_exc()
        messagebox.showerror("Erro ao Salvar", f"Ocorreu um erro ao salvar o arquivo:\n{e}")
        return False

def get_todos_logs_filtrados():
    caminho_busca = os.path.join(_base_dir(), PASTA_LOGS_EXCEL, "**", "*.xlsx")
    todos_arquivos_brutos = glob.glob(caminho_busca, recursive=True)
    arquivos_validos = [f for f in todos_arquivos_brutos if not os.path.basename(f).startswith('~')]
    return sorted(arquivos_validos, reverse=True)

def get_containers_no_patio():
    todos_arquivos = get_todos_logs_filtrados()
    if not todos_arquivos:
        return pd.DataFrame()
    lista_dfs = []
    for f in todos_arquivos:
        if os.path.getsize(f) > 0:
            try:
                df = pd.read_excel(f, sheet_name=NOME_ABA_EXCEL)
                for col in df.columns:
                    if col != 'Data e Hora':
                        df[col] = df[col].astype(str).replace('nan', '')
                lista_dfs.append(df)
            except Exception:
                continue
    if not lista_dfs:
        return pd.DataFrame()
    df_total = pd.concat(lista_dfs, ignore_index=True)
    df_total['Data e Hora'] = pd.to_datetime(df_total['Data e Hora'], errors='coerce')
    df_total.dropna(subset=['Data e Hora', 'Nº do Contêiner'], inplace=True)
    df_total.sort_values(by='Data e Hora', ascending=False, inplace=True)
    ultimo_movimento = df_total.drop_duplicates(subset='Nº do Contêiner', keep='first')
    return ultimo_movimento[ultimo_movimento['Status'] == 'Entrada']

# --- UI HELPERS ---
def formatar_cpf_aprimorado(event, entry_widget):
    posicao_cursor = entry_widget.index(INSERT)
    texto_antes = entry_widget.get()
    entry_widget.unbind("<KeyRelease>")
    numeros = ''.join(filter(str.isdigit, texto_antes))[:11]
    if len(numeros) > 9:
        texto_formatado = f"{numeros[:3]}.{numeros[3:6]}.{numeros[6:9]}-{numeros[9:]}"
    elif len(numeros) > 6:
        texto_formatado = f"{numeros[:3]}.{numeros[3:6]}.{numeros[6:]}"
    elif len(numeros) > 3:
        texto_formatado = f"{numeros[:3]}.{numeros[3:]}"
    else:
        texto_formatado = numeros
    entry_widget.delete(0, END)
    entry_widget.insert(0, texto_formatado)
    if len(texto_formatado) > len(texto_antes):
        entry_widget.icursor(END)
    else:
        entry_widget.icursor(posicao_cursor)
    entry_widget.bind("<KeyRelease>", lambda e: formatar_cpf_aprimorado(e, entry_widget))

def formatar_texto_maiusculo(event, entry_widget):
    posicao_cursor = entry_widget.index(INSERT)
    texto_atual = entry_widget.get()
    entry_widget.delete(0, END)
    entry_widget.insert(0, texto_atual.upper())
    entry_widget.icursor(posicao_cursor)

def formatar_cpf_para_exibicao(cpf_numeros):
    cpf_numeros_str = ''.join(filter(str.isdigit, str(cpf_numeros)))
    if len(cpf_numeros_str) != 11:
        return cpf_numeros_str
    return f"{cpf_numeros_str[:3]}.{cpf_numeros_str[3:6]}.{cpf_numeros_str[6:9]}-{cpf_numeros_str[9:]}"

def limpar_campos():
    entry_container.delete(0, END)
    entry_placa.delete(0, END)
    entry_motorista.delete(0, END)
    entry_cpf.delete(0, END)
    entry_cliente.delete(0, END)
    tipo_container_var.set(TIPOS_CONTAINER[0])
    condicao_var.set(CONDICOES[0])
    entry_lacre.delete(0, END)
    entry_nf.delete(0, END)
    entry_destino.delete(0, END)
    entry_obs.delete("1.0", END)
    entry_container.focus_set()

# --- JANELAS ---
def abrir_janela_patio():
    patio_window = ttk.Toplevel(title="Controle de Pátio - Contêineres Presentes")
    patio_window.geometry("800x500")
    patio_window.transient(app)
    patio_window.grab_set()

    def carregar_dados_patio():
        for item in tree.get_children():
            tree.delete(item)
        try:
            containers_no_patio = get_containers_no_patio()
            if containers_no_patio.empty:
                messagebox.showinfo("Pátio Vazio", "Nenhum contêiner encontrado no pátio.", parent=patio_window)
                return
            df_para_exibir = containers_no_patio.copy()
            df_para_exibir['Data e Hora'] = df_para_exibir['Data e Hora'].dt.strftime('%d/%m/%Y %H:%M')
            for _, row in df_para_exibir.iterrows():
                tree.insert("", END, values=(row.get('Nº do Contêiner', ''), row.get('Cliente', ''), row.get('Placa do Veículo', ''), row.get('Data e Hora', '')))
        except Exception as e:
            messagebox.showerror("Erro ao Ler Histórico", f"Ocorreu um erro:\n{e}", parent=patio_window)

    ttk.Button(patio_window, text="Atualizar Lista", command=carregar_dados_patio, bootstyle="info").pack(pady=10)
    cols = ('Nº do Contêiner', 'Cliente', 'Placa do Veículo', 'Data de Entrada')
    tree = ttk.Treeview(patio_window, columns=cols, show='headings')
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=180)
    tree.pack(expand=True, fill=BOTH, padx=10, pady=5)
    carregar_dados_patio()

def abrir_janela_saida():
    selecao_window = ttk.Toplevel(title="Registrar Saída - Etapa 1 de 2")
    selecao_window.geometry("800x500")
    selecao_window.transient(app)
    selecao_window.grab_set()

    def carregar_containers_patio():
        for item in tree.get_children():
            tree.delete(item)
        try:
            df_patio = get_containers_no_patio()
            if df_patio.empty:
                info_label.config(text="Nenhum contêiner no pátio.")
            else:
                info_label.config(text="Selecione o contêiner que está saindo:")
            for index, row in df_patio.iterrows():
                tree.insert("", END, values=[row.get(col, '') for col in COLUNAS_ORDENADAS], tags=(index,))
        except Exception as e:
            messagebox.showerror("Erro ao Ler Histórico", f"Ocorreu um erro:\n{e}", parent=selecao_window)

    def on_selecionar_container():
        selecionado = tree.focus()
        if not selecionado:
            messagebox.showwarning("Atenção", "Nenhum contêiner foi selecionado.", parent=selecao_window)
            return
        index_selecionado = tree.item(selecionado)['tags'][0]
        dados_originais = get_containers_no_patio().loc[index_selecionado].to_dict()
        selecao_window.destroy()
        abrir_janela_dados_saida_e_romaneio(dados_originais)

    info_label = ttk.Label(selecao_window, text="Carregando...", font=("-size 12 -weight bold"))
    info_label.pack(pady=(10,5))

    cols_visiveis = ('Nº do Contêiner', 'Cliente', 'Placa do Veículo', 'Data e Hora')
    tree = ttk.Treeview(selecao_window, columns=COLUNAS_ORDENADAS, displaycolumns=cols_visiveis, show='headings')
    for col in cols_visiveis:
        tree.heading(col, text=col)
        tree.column(col, width=180)
    tree.pack(expand=True, fill=BOTH, padx=10, pady=5)

    ttk.Button(selecao_window, text="Próximo -> Preencher Dados de Saída", command=on_selecionar_container, bootstyle="primary", padding=10).pack(pady=10)
    carregar_containers_patio()

def abrir_janela_dados_saida_e_romaneio(dados_container):
    dialog = ttk.Toplevel(title="Registrar Saída - Etapa 2 de 2")
    dialog.geometry("850x600")
    dialog.transient(app)
    dialog.grab_set()
    entries = {}

    frame_info = ttk.LabelFrame(dialog, text="Dados do Contêiner (Entrada)", padding=10)
    frame_info.pack(padx=10, pady=10, fill=X)

    frame_transporte = ttk.LabelFrame(dialog, text="1. Informações do Transporte de Saída", padding=10)
    frame_transporte.pack(padx=10, pady=10, fill=X)

    frame_romaneio = ttk.LabelFrame(dialog, text="2. Informações do Romaneio", padding=10)
    frame_romaneio.pack(padx=10, pady=10, fill=X)

    info_text = f"Contêiner: {dados_container['Nº do Contêiner']}  |  Cliente: {dados_container['Cliente']}  |  Entrada: {dados_container['Data e Hora'].strftime('%d/%m/%Y %H:%M')}"
    ttk.Label(frame_info, text=info_text, font=("-weight bold")).pack()

    campos_transporte = ['Placa do Veículo', 'Placa Carreta', 'Motorista', 'CPF Motorista', 'Transportadora']
    for i, campo in enumerate(campos_transporte):
        ttk.Label(frame_transporte, text=f"{campo}:").grid(row=i, column=0, padx=5, pady=5, sticky='w')
        entries[campo] = ttk.Entry(frame_transporte, width=40)
        entries[campo].grid(row=i, column=1, padx=5, pady=5, sticky='w')
        if "Placa" in campo:
            entries[campo].bind("<KeyRelease>", lambda e, w=entries[campo]: formatar_texto_maiusculo(e, w))
        if "CPF" in campo:
            entries[campo].bind("<KeyRelease>", lambda e, w=entries[campo]: formatar_cpf_aprimorado(e, w))

    campos_romaneio = ['Tara', 'Peso Bruto Carga', 'Booking', 'Armador', 'Navio', 'Deadline']
    for i, campo in enumerate(campos_romaneio):
        ttk.Label(frame_romaneio, text=f"{campo}:").grid(row=i % 3, column=(i // 3) * 2, padx=5, pady=5, sticky='w')
        entries[campo] = ttk.Entry(frame_romaneio, width=30)
        entries[campo].grid(row=i % 3, column=(i // 3) * 2 + 1, padx=5, pady=5, sticky='w')

    def on_confirmar_tudo():
        dados_novos = {campo: entries[campo].get().strip() for campo in campos_transporte + campos_romaneio}

        if not dados_novos['Placa do Veículo'] or not dados_novos['Motorista'] or not dados_novos['Transportadora']:
            messagebox.showwarning("Atenção", "Placa do Veículo, Motorista e Transportadora são obrigatórios.", parent=dialog)
            return

        cpf_limpo = ''.join(filter(str.isdigit, dados_novos['CPF Motorista']))
        if cpf_limpo and not validar_cpf(cpf_limpo):
            messagebox.showerror("CPF Inválido", f"O CPF '{dados_novos['CPF Motorista']}' é inválido.", parent=dialog)
            return

        if not all(dados_novos[k] for k in campos_romaneio):
            messagebox.showwarning("Atenção", "Todos os campos do Romaneio devem ser preenchidos.", parent=dialog)
            return

        try:
            # localizar log da entrada original
            todos_logs = get_todos_logs_filtrados()
            arquivo_do_registro = None
            for log_file in todos_logs:
                df_temp = pd.read_excel(log_file, sheet_name=NOME_ABA_EXCEL)
                df_temp['Data e Hora'] = pd.to_datetime(df_temp['Data e Hora'])
                if dados_container['Data e Hora'] in df_temp['Data e Hora'].values:
                    arquivo_do_registro = log_file
                    break

            if not arquivo_do_registro:
                messagebox.showerror("Erro Crítico", "Não foi possível encontrar o arquivo de log original da entrada.", parent=dialog)
                return

            df_log = pd.read_excel(arquivo_do_registro, sheet_name=NOME_ABA_EXCEL)
            for col in df_log.columns:
                if col != 'Data e Hora':
                    df_log[col] = df_log[col].astype(str).replace('nan', '')
            df_log['Data e Hora'] = pd.to_datetime(df_log['Data e Hora'])

            indice_para_atualizar = df_log[df_log['Data e Hora'] == dados_container['Data e Hora']].index
            if indice_para_atualizar.empty:
                messagebox.showerror("Erro Crítico", "Não foi possível encontrar a linha de entrada original no log.", parent=dialog)
                return

            for campo_romaneio in campos_romaneio:
                df_log.loc[indice_para_atualizar, campo_romaneio] = dados_novos[campo_romaneio]

            dados_finais_saida = df_log.loc[indice_para_atualizar].to_dict('records')[0]
            dados_finais_saida.update(dados_novos)
            dados_finais_saida['CPF Motorista'] = cpf_limpo
            dados_finais_saida['Status'] = 'Saída'
            dados_finais_saida['Data e Hora'] = datetime.now()

            df_log = pd.concat([df_log, pd.DataFrame([dados_finais_saida])], ignore_index=True)
            salvar_planilha(df_log, arquivo_do_registro)

            # PDF do romaneio
            os.makedirs(os.path.join(_base_dir(), PASTA_ROMANEIOS_PDF), exist_ok=True)
            nome_pdf = f"ROMANEIO_{dados_finais_saida['Nº do Contêiner'].replace('/', '-')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            caminho_pdf = os.path.join(_base_dir(), PASTA_ROMANEIOS_PDF, nome_pdf)

            pdf = FPDF(); pdf.add_page(); pdf.set_auto_page_break(auto=True, margin=15)
            pdf.set_font("Helvetica", "B", 16)
            pdf.cell(0, 10, "Romaneio de Transporte", new_x="LMARGIN", new_y="NEXT", align="C")
            pdf.ln(10)

            dados_pdf = {
                'Nº do Contêiner': dados_finais_saida['Nº do Contêiner'], 'Cliente': dados_finais_saida['Cliente'],
                'Placa do Veículo': dados_finais_saida['Placa do Veículo'], 'Placa Carreta': dados_finais_saida['Placa Carreta'],
                'Motorista': dados_finais_saida['Motorista'], 'CPF': formatar_cpf_para_exibicao(dados_finais_saida['CPF Motorista']),
                'Transportadora': dados_finais_saida['Transportadora'], 'Nº do Lacre': dados_finais_saida['Nº do Lacre'],
                'Tara': dados_finais_saida['Tara'], 'Peso da Carga': dados_finais_saida['Peso Bruto Carga'],
                'Booking': dados_finais_saida['Booking'], 'Armador': dados_finais_saida['Armador'],
                'Navio': dados_finais_saida['Navio'], 'Deadline': dados_finais_saida['Deadline'],
                'Destino': dados_finais_saida['Destino']
            }

            for chave, valor in dados_pdf.items():
                pdf.set_font("Helvetica", "B", 12); pdf.cell(45, 10, f"{chave}:", border=0)
                pdf.set_font("Helvetica", "", 12); pdf.multi_cell(0, 10, str(valor), border=0, new_x="LMARGIN", new_y="NEXT")

            pdf.ln(20); pdf.cell(0, 10, "________________________________________", new_x="LMARGIN", new_y="NEXT", align="C")
            pdf.cell(0, 10, "Assinatura Motorista", new_x="LMARGIN", new_y="NEXT", align="C")
            pdf.output(caminho_pdf)

            messagebox.showinfo("Sucesso", f"Saída registrada e Romaneio gerado!\nSalvo em: {caminho_pdf}", parent=dialog)
            dialog.destroy()
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Erro Detalhado", f"Ocorreu um erro completo no processo de saída:\n\n{e}\n\nVerifique o terminal para mais detalhes.", parent=dialog)

    ttk.Button(dialog, text="Confirmar Saída e Gerar Documentos", command=on_confirmar_tudo, bootstyle="success", padding=10).pack(pady=20)

# --- INTERFACE PRINCIPAL ---
app = ttk.Window(themename="darkly")
app.title("M4 Logística - Controle de Movimentação")
app.geometry("980x560")

main_frame = ttk.Frame(app, padding=(20, 10))
main_frame.pack(fill=BOTH, expand=True)

ttk.Label(main_frame, text="Nº do Contêiner:", font=("-weight bold")).grid(row=0, column=0, padx=10, pady=8, sticky='w')
entry_container = ttk.Entry(main_frame, width=30)
entry_container.grid(row=0, column=1, padx=10, pady=8)
entry_container.bind("<KeyRelease>", lambda e: formatar_texto_maiusculo(e, entry_container))

ttk.Label(main_frame, text="Placa do Veículo:", font=("-weight bold")).grid(row=1, column=0, padx=10, pady=8, sticky='w')
entry_placa = ttk.Entry(main_frame, width=30)
entry_placa.grid(row=1, column=1, padx=10, pady=8)
entry_placa.bind("<KeyRelease>", lambda e: formatar_texto_maiusculo(e, entry_placa))

ttk.Label(main_frame, text="Motorista:").grid(row=2, column=0, padx=10, pady=8, sticky='w')
entry_motorista = ttk.Entry(main_frame, width=30)
entry_motorista.grid(row=2, column=1, padx=10, pady=8)

ttk.Label(main_frame, text="CPF Motorista:").grid(row=3, column=0, padx=10, pady=8, sticky='w')
entry_cpf = ttk.Entry(main_frame, width=30)
entry_cpf.grid(row=3, column=1, padx=10, pady=8)
entry_cpf.bind("<KeyRelease>", lambda e: formatar_cpf_aprimorado(e, entry_cpf))

ttk.Label(main_frame, text="Cliente:").grid(row=4, column=0, padx=10, pady=8, sticky='w')
entry_cliente = ttk.Entry(main_frame, width=30)
entry_cliente.grid(row=4, column=1, padx=10, pady=8)

ttk.Label(main_frame, text="Nº do Lacre:").grid(row=5, column=0, padx=10, pady=8, sticky='w')
entry_lacre = ttk.Entry(main_frame, width=30)
entry_lacre.grid(row=5, column=1, padx=10, pady=8)

ttk.Label(main_frame, text="Tipo de Contêiner:").grid(row=0, column=2, padx=15, pady=8, sticky='w')
tipo_container_var = ttk.StringVar(value=TIPOS_CONTAINER[0])
combo_tipo = ttk.Combobox(main_frame, textvariable=tipo_container_var, values=TIPOS_CONTAINER, width=28, state="readonly")
combo_tipo.grid(row=0, column=3, padx=10, pady=8)

ttk.Label(main_frame, text="Condição:").grid(row=1, column=2, padx=15, pady=8, sticky='w')
condicao_var = ttk.StringVar(value=CONDICOES[0])
combo_condicao = ttk.Combobox(main_frame, textvariable=condicao_var, values=CONDICOES, width=28, state="readonly")
combo_condicao.grid(row=1, column=3, padx=10, pady=8)

ttk.Label(main_frame, text="Nota Fiscal (NF):").grid(row=2, column=2, padx=15, pady=8, sticky='w')
entry_nf = ttk.Entry(main_frame, width=30)
entry_nf.grid(row=2, column=3, padx=10, pady=8)

ttk.Label(main_frame, text="Destino:").grid(row=3, column=2, padx=15, pady=8, sticky='w')
entry_destino = ttk.Entry(main_frame, width=30)
entry_destino.grid(row=3, column=3, padx=10, pady=8)

ttk.Label(main_frame, text="Observações:").grid(row=6, column=0, padx=10, pady=8, sticky='w')
entry_obs = ttk.Text(main_frame, height=4, width=90)
entry_obs.grid(row=7, column=0, columnspan=4, padx=10, pady=8)

def registrar_entrada_wrapper():
    dados_entrada = {
        'Nº do Contêiner': entry_container.get().strip().upper(),
        'Placa do Veículo': entry_placa.get().strip().upper(),
        'Motorista': entry_motorista.get().strip(),
        'CPF Motorista': ''.join(filter(str.isdigit, entry_cpf.get())),
        'Cliente': entry_cliente.get().strip(),
        'Tipo de Contêiner': tipo_container_var.get(),
        'Condição': condicao_var.get(),
        'Nº do Lacre': entry_lacre.get().strip(),
        'Nota Fiscal (NF)': entry_nf.get().strip(),
        'Destino': entry_destino.get().strip(),
        'Observações': entry_obs.get("1.0", "end-1c").strip(),
        'Data e Hora': datetime.now(),
        'Status': 'Entrada'
    }
    if not dados_entrada['Nº do Contêiner'] or not dados_entrada['Placa do Veículo']:
        messagebox.showwarning("Atenção", "Os campos 'Nº do Contêiner' e 'Placa do Veículo' são obrigatórios.")
        return
    if not validar_numero_container(dados_entrada['Nº do Contêiner']):
        messagebox.showerror("Erro de Validação", f"O número de contêiner '{dados_entrada['Nº do Contêiner']}' é INVÁLIDO!")
        return
    if len(dados_entrada['Placa do Veículo']) < 7:
        messagebox.showwarning("Validação", "A Placa do Veículo deve ter no mínimo 7 caracteres.")
        return
    if dados_entrada['CPF Motorista'] and not validar_cpf(dados_entrada['CPF Motorista']):
        messagebox.showerror("Erro de Validação", f"O CPF '{entry_cpf.get()}' é INVÁLIDO.")
        return
    for col in ['Placa Carreta', 'Transportadora', 'Tara', 'Peso Bruto Carga', 'Booking', 'Armador', 'Navio', 'Deadline']:
        dados_entrada[col] = ''
    if registrar_movimento('Entrada', dados_entrada):
        messagebox.showinfo("Sucesso", "Registro de ENTRADA salvo com sucesso!")
        limpar_campos()

# BOTÕES
frame_botoes = ttk.Frame(main_frame)
frame_botoes.grid(row=8, column=0, columnspan=4, pady=20)

ttk.Button(frame_botoes, text="Registrar Entrada", width=20, bootstyle="primary", command=registrar_entrada_wrapper, padding=(10, 10)).pack(side=LEFT, padx=5)
ttk.Button(frame_botoes, text="Registrar Saída", width=20, bootstyle="danger", command=abrir_janela_saida, padding=(10, 10)).pack(side=LEFT, padx=5)
ttk.Button(frame_botoes, text="Ver Pátio Atual", width=20, bootstyle="success", command=abrir_janela_patio, padding=(10, 10)).pack(side=LEFT, padx=5)

def abrir_excel_do_dia():
    caminho = garantir_arquivo_do_dia()
    _abrir_no_sistema(caminho)

def abrir_pasta_logs_dia():
    pasta = get_pasta_logs_do_dia()
    os.makedirs(pasta, exist_ok=True)
    _abrir_no_sistema(pasta)

ttk.Button(frame_botoes, text="Abrir Excel do Dia", width=20, bootstyle="secondary", command=abrir_excel_do_dia, padding=(10,10)).pack(side=LEFT, padx=5)
ttk.Button(frame_botoes, text="Abrir Pasta dos Logs", width=20, bootstyle="secondary", command=abrir_pasta_logs_dia, padding=(10,10)).pack(side=LEFT, padx=5)
#teste
if __name__ == "__main__":
    app.mainloop()
